from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q, Count, Case, When, Value, IntegerField, Avg
from datetime import date
from .models import DemandaETL
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .forms import DemandaETLForm


def home(request):
    busca = request.GET.get('search') or request.GET.get('q')
    ver_producao = request.GET.get('ver_producao') == 'on'
    hoje = date.today()

    # 1. QuerySet inicial
    demandas_queryset = DemandaETL.objects.all()

    # 2. Aplicar Filtro de Produção (Ocultar por padrão)
    if not ver_producao:
        demandas_queryset = demandas_queryset.exclude(status='P')

    # 3. Aplicar busca por texto
    if busca:
        demandas_queryset = demandas_queryset.filter(
            Q(titulo__icontains=busca) |
            Q(id_demanda__icontains=busca) |
            Q(workflow_mapping__icontains=busca) |
            Q(folder_repositorio__icontains=busca)
        )

    demandas_concluidas = DemandaETL.objects.filter(status='P').exclude(data_implementacao__isnull=True)
    soma_dias = 0
    count_concluidas = demandas_concluidas.count()

    for d in demandas_concluidas:
        delta = d.data_implementacao - d.data_recebimento
        soma_dias += max(0, delta.days)

    media_lead_time = round(soma_dias / count_concluidas, 1) if count_concluidas > 0 else 0



    # 4. Loop ÚNICO para processar Semáforo, Lead Time e Pesos
    for d in demandas_queryset:
        # Cálculo de Semáforo e Peso de Prioridade
        if d.data_implementacao:
            dias_restantes = (d.data_implementacao - hoje).days
            if dias_restantes < 0:
                d.semaforo_cor = "danger"
                d.semaforo_msg = f"Atrasada ({abs(dias_restantes)}d)"
                d.peso_prioridade = 1
            elif dias_restantes <= 3:
                d.semaforo_cor = "warning"
                d.semaforo_msg = "Urgente"
                d.peso_prioridade = 2
            else:
                d.semaforo_cor = "success"
                d.semaforo_msg = "No Prazo"
                d.peso_prioridade = 3
        else:
            d.semaforo_cor = "secondary"
            d.semaforo_msg = "Sem Data"
            d.peso_prioridade = 4

        # Cálculo de Lead Time (Tempo de Execução)
        if d.status == 'P' and d.data_implementacao:
            delta = d.data_implementacao - d.data_recebimento
            d.tempo_execucao = max(0, delta.days)
        else:
            d.tempo_execucao = None

        # Lógica de Porcentagem para a Barra de Progresso
        if d.status == 'D':
            d.progresso = 30
            d.progresso_cor = "warning"
        elif d.status == 'T':
            d.progresso = 70
            d.progresso_cor = "info"
        elif d.status == 'P':
            d.progresso = 100
            d.progresso_cor = "success"
        else:
            d.progresso = 0
            d.progresso_cor = "secondary"

    # 5. Ordenação Segura (Garante que todas tenham o atributo)
    demandas_final = sorted(demandas_queryset, key=lambda x: getattr(x, 'peso_prioridade', 4))

    # 6. Estatísticas para os Cards (Baseadas no total geral do banco)
    # Usamos .all() aqui para os números do topo não sumirem ao filtrar a lista
    total_geral = DemandaETL.objects.count()
    em_desenv = DemandaETL.objects.filter(status='D').count()
    alta = DemandaETL.objects.filter(complexidade='A').count()

    # 7. Dados para o Gráfico
    stats_status = DemandaETL.objects.values('status').annotate(total=Count('status'))
    labels = []
    data_grafico = []
    status_map = dict(DemandaETL.STATUS_CHOICES)
    for s in stats_status:
        labels.append(status_map.get(s['status']))
        data_grafico.append(s['total'])

    context = {
        'demandas': demandas_final,
        'total': total_geral,
        'em_desenv': em_desenv,
        'alta': alta,
        'valor_busca': busca,
        'ver_producao': ver_producao,
        'labels': labels,
        'data_grafico': data_grafico,
        'media_lead_time': media_lead_time,
    }
    return render(request, 'demandas/home.html', context)


# --- FUNÇÃO DE EXPORTAÇÃO ---
def exportar_excel(request):
    busca = request.GET.get('search')
    demandas_queryset = DemandaETL.objects.all().order_by('-data_recebimento')

    if busca:
        demandas_queryset = demandas_queryset.filter(
            Q(titulo__icontains=busca) | Q(id_demanda__icontains=busca)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio_ETL"

    # Estilos profissionais para o Excel
    header_font = Font(name='Arial', bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000851", end_color="000851", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabeçalho da Planilha
    colunas = ['ID Demanda', 'Título', 'Status', 'Complexidade', 'TL Responsável', 'Link Jira', 'Pasta PC']
    ws.append(colunas)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Inserção dos Dados
    for d in demandas_queryset:
        ws.append([
            d.id_demanda,
            d.titulo,
            d.get_status_display(),
            d.get_complexidade_display(),
            d.lider_tecnico,
            d.link_jira or 'N/A',
            d.folder_repositorio
        ])

    # Ajuste automático de largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    ws.auto_filter.ref = ws.dimensions

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Relatorio_Demandas_ETL.xlsx'
    wb.save(response)

    return response

def alterar_status(request, id):
    """Altera o status da demanda"""
    demanda = get_object_or_404(DemandaETL, id=id)

    if request.method == 'POST':
        novo_status = request.POST.get('status')

        if novo_status in ['D', 'T', 'P']:
            demanda.status = novo_status
            demanda.save()

            messages.success(
                request,
                f'Status alterado para {demanda.get_status_display()} com sucesso!'
            )
        else:
            messages.error(request, 'Status inválido.')
    return redirect('home')

def excluir_demanda(request, id):
    demanda = get_object_or_404(DemandaETL, id=id)

    if request.method == "POST":
        demanda.delete()
        messages.success(request, "Demanda excluída com sucesso!")

    return redirect('home')

def editar_demanda(request, id):
    demanda = get_object_or_404(DemandaETL, id=id)

    if request.method == 'POST':
        form = DemandaETLForm(request.POST, instance=demanda)
        if form.is_valid():
            form.save()
            messages.success(request, "Demanda atualizada com sucesso!")
            return redirect('home')
    else:
        form = DemandaETLForm(instance=demanda)

    return render(request, 'demandas/editar_demanda.html', {
        'form': form,
        'demanda': demanda
    })

def deletar_demanda(request, id):
    demanda = get_object_or_404(DemandaETL, id=id)

    if request.method == "POST":
        titulo = demanda.titulo
        demanda.delete()
        messages.success(request, f"Demanda '{titulo}' excluída com sucesso!")

    return redirect('home')